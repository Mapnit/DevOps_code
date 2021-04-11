import argparse
import copy
import os
from datetime import datetime

import pytz
from azure.devops.connection import Connection
from azure.devops.v5_1.work import models as workModels
from azure.devops.v5_1.work_item_tracking import \
    models as workItemTrackingModels
from msrest.authentication import BasicAuthentication
from openpyxl import Workbook
from openpyxl import load_workbook

# DevOps personal access token (recycle every 30 days) 
personal_access_token = 'urlycbelh7lf4bc6w3c3a4jd2xur3uwobfpahbswyzekcy2mucaq'
# DevOps Organization URL
organization_url = 'https://dev.azure.com/CenterPoint-GIS'

usa_cst = pytz.timezone('US/Central')
today_timezone = usa_cst.localize(datetime.now())

# Create a connection to the org
credentials = BasicAuthentication('', personal_access_token)
connection = Connection(base_url=organization_url, creds=credentials)

class experiment: 
    def get_projects():
        # Get a client (the "core" client provides access to projects, teams, etc)
        core_client = connection.clients.get_core_client()

        # Get the first page of projects
        get_projects_response = core_client.get_projects()

        index = 0
        if get_projects_response is not None:
            for project in get_projects_response.value:
                print("Project [" + str(index) + "]: " + project.name)
                index += 1
            if get_projects_response.continuation_token is not None and get_projects_response.continuation_token != "":
                # Get the next page of projects
                get_projects_response = core_client.get_projects(continuation_token=get_projects_response.continuation_token)
            else:
                # All projects have been retrieved
                get_projects_response = None

    def get_teams():
        project_name = 'CNP.GIS'
        # Get a client (the "core" client provides access to projects, teams, etc)
        core_client = connection.clients.get_core_client()
        # get the teams
        index = 0
        get_teams_response = core_client.get_teams(project_name) 
        if get_teams_response is not None: 
            for team in get_teams_response: 
                print("Team [" + str(index) + "]: " + team.name)
                index += 1
            # All teams have been retrieved
            get_teams_response = None

    def get_boards(): 
        # get the boards 
        index = 0 
        get_boards_response = work_client.get_boards(team_context)

        if get_boards_response is not None: 
            for board in get_boards_response: 
                print("Board [" + str(index) + "]: " + board.name)
                index += 1
            # All boards have been retrieved
            get_teams_response = None

work_item_type_of_interest = ["Product Backlog Item", "Task"]

pbi_wiql_template = "\
    Select [System.Id] From WorkItems \
    Where [System.AreaPath] = '{AreaPath}' \
        and [System.WorkItemType] in ('Product Backlog Item') \
        and [System.State] <> 'Removed' \
        and [System.IterationPath] = '{IterationPath}' \
    Order by [Microsoft.VSTS.Common.Priority] asc, [System.CreatedDate] desc \
"

task_wiql_template = "\
    Select [System.Id] From WorkItems \
    Where [System.AreaPath] = '{AreaPath}' \
        and [System.WorkItemType] in ('Task') \
        and [System.State] <> 'Removed' \
        and [System.IterationPath] = '{IterationPath}' \
    Order by [Microsoft.VSTS.Common.Priority] asc, [System.CreatedDate] desc \
"

pbi_field_names = ['System.Id', 'System.WorkItemType', 'System.Parent', 
    'System.Title', 'System.Tags', # 'System.Description',
    'Microsoft.VSTS.Common.ValueArea', 'Microsoft.VSTS.Common.BusinessValue', 
    'System.AssignedTo', 'System.State', 'System.CreatedDate', 'System.ChangedDate',
    'System.AreaPath', 'System.IterationPath']

# two weeks long 
DAYS_PER_ITERATION = 10 

capacity_field_names = ['team_member', 'capacity_per_day', 
    'days_per_iteration', 'IterationPath']


def get_current_iteration(team_context): 

    work_client = connection.clients.get_work_client()

    iteration_id = None
    iteration_path = None
    iteration_due_date = None

    # get the iteration paths
    index = 0 
    get_team_iterations_response = work_client.get_team_iterations(team_context, 'current')
    if get_team_iterations_response is not None: 
        for team_iteration in get_team_iterations_response: 
            start_date = team_iteration.attributes.start_date
            if start_date.year < 2021 or start_date > today_timezone: 
                # skip the iterations of last year and the future
                None
            else: 
                iteration_path = team_iteration.path
                iteration_id = team_iteration.id
                iteration_due_date = team_iteration.attributes.finish_date 
                print("Iteration [{0}]: {1}, {2} ({3} -> {4})".format(
                    index, team_iteration.name, team_iteration.path, 
                    team_iteration.attributes.start_date.strftime("%Y-%m-%d"),
                    team_iteration.attributes.finish_date.strftime("%Y-%m-%d")
                    ))
                index += 1
        # All team iterations have been retrieved
        get_team_iterations_response = None

    return iteration_id, current_iteration, iteration_due_date


def get_iteration(team_context, iteration_path): 

    work_client = connection.clients.get_work_client()

    iteration_id = None
    iteration_due_date = None

    # get the iteration paths
    index = 0 
    get_team_iterations_response = work_client.get_team_iterations(team_context)
    if get_team_iterations_response is not None: 
        for team_iteration in get_team_iterations_response: 
            start_date = team_iteration.attributes.start_date
            if start_date.year < 2021 or start_date > today_timezone: 
                # skip the iterations of last year and the future
                continue
            elif iteration_path == team_iteration.path:
                iteration_id = team_iteration.id
                iteration_due_date = team_iteration.attributes.finish_date 
                print("Iteration [{0}]: {1}, {2} ({3} -> {4})".format(
                    index, team_iteration.name, team_iteration.path, 
                    team_iteration.attributes.start_date.strftime("%Y-%m-%d"),
                    team_iteration.attributes.finish_date.strftime("%Y-%m-%d")
                    ))
                index += 1
        # All team iterations have been retrieved
        get_team_iterations_response = None

    return iteration_id, iteration_due_date


def get_past_iterations(team_context): 

    work_client = connection.clients.get_work_client()

    iteration_list = []

    # get the iteration paths
    index = 0 
    get_team_iterations_response = work_client.get_team_iterations(team_context)
    if get_team_iterations_response is not None: 
        for team_iteration in get_team_iterations_response: 
            start_date = team_iteration.attributes.start_date
            if start_date.year < 2021 or start_date > today_timezone: 
                # skip the iterations of last year and the future
                continue
            else:
                iteration_path = team_iteration.path
                iteration_id = team_iteration.id
                iteration_due_date = team_iteration.attributes.finish_date                 
                print("Iteration [{0}]: {1}, {2} ({3} -> {4})".format(
                    index, team_iteration.name, team_iteration.path, 
                    team_iteration.attributes.start_date.strftime("%Y-%m-%d"),
                    team_iteration.attributes.finish_date.strftime("%Y-%m-%d")
                    ))
                iteration_list.append({
                    'iteration_id': iteration_id, 
                    'iteration_path': iteration_path, 
                    'iteration_due_date': iteration_due_date
                })
                index += 1
        # All team iterations have been retrieved
        get_team_iterations_response = None

    return iteration_list


def _retrieve_work_items(team_context, wiql_query):

    # query the backlogs 
    work_tracking_client = connection.clients.get_work_item_tracking_client()

    wiql = workItemTrackingModels.Wiql(query=wiql_query)

    query_by_wiql_response = work_tracking_client.query_by_wiql(wiql, team_context)

    index = 0
    idList = []
    parentIdList = []
    work_item_list = []
    if query_by_wiql_response is not None:
        # collect all work item Ids 
        for work_item_id in query_by_wiql_response.work_items: 
            # print("work item [{0}]: {1}".format(index, work_item_id.id))
            index += 1
            idList.append(work_item_id.id)

        # output all work items, including parents not created in this iteration
        i, j = 0, min(200, len(idList))
        while i < len(idList):    
            # get work items in a batch 
            get_work_items_response = work_tracking_client.get_work_items(
                idList[i:j], fields = pbi_field_names)
            if get_work_items_response is not None: 
                for work_item in get_work_items_response: 
                    # output to the screen
                    print("{0}, {1}: {2}".format(
                        work_item.fields["System.WorkItemType"], 
                        work_item.fields["System.Title"], 
                        work_item.fields["System.State"]))
                    # save to the return variable
                    work_item_list.append(work_item)

            i, j = j, min(j + 200, len(idList))

        # All query results have been retrieved
        print("total {0} work items retrieved".format(index))
        idList = None
        query_by_wiql_response = None

        # return all work items
        return work_item_list


def retrieve_PBIs(team_context, iteration_path):

    wiql_pbi_query = pbi_wiql_template.replace( \
        "{AreaPath}", team_context.project).replace( \
        "{IterationPath}", iteration_path)

    return _retrieve_work_items(team_context, wiql_pbi_query)


def retrieve_tasks(team_context, iteration_path): 

    wiql_task_query = task_wiql_template.replace( \
        "{AreaPath}", team_context.project).replace( \
        "{IterationPath}", iteration_path)
        #"{ParentId}", str(pbi_id)) # System.Parent not applicable in filter 

    return _retrieve_work_items(team_context, wiql_task_query)


def get_lead_duration(team_context, item_id): 
    # query the revision 
    work_tracking_client = connection.clients.get_work_item_tracking_client()

    get_updates_response =  work_tracking_client.get_updates(item_id, project=team_context.project)

    create_date = None
    start_date = None
    finish_date = None
    if get_updates_response is not None:
        for revision in get_updates_response: 
            if revision.fields is None:
                # skip any non-field update
                continue
            else: 
                field_revision = revision.fields
                fields_of_interest = ['System.State']
                for field_name in fields_of_interest:
                    if field_name in field_revision.keys(): 
                        iteration_path = None
                        if 'System.IterationPath' in field_revision.keys(): 
                            iteration_path = field_revision['System.IterationPath'].new_value
                        changed_date = field_revision['System.ChangedDate'].new_value
                        print("{0} - {1} [{2}, {3}]: {4} -> {5}".format(item_id, field_name, 
                            changed_date, iteration_path, 
                            field_revision[field_name].old_value, field_revision[field_name].new_value))
                        if field_revision[field_name].new_value == 'New' and create_date is None: 
                            # the first start date of Product Backlog Item
                            create_date = changed_date
                            # new work found. reset the finish date 
                            finish_date = None
                        elif field_revision[field_name].new_value == 'To Do' and create_date is None: 
                            # the first start date of Task
                            create_date = changed_date
                            # new work found. reset the finish date 
                            finish_date = None
                        elif field_revision[field_name].new_value == 'Started' and start_date is None: 
                            # the first start date of Product Backlog Item
                            start_date = changed_date
                            # new work found. reset the finish date 
                            finish_date = None
                        elif field_revision[field_name].new_value == 'In Progress' and start_date is None: 
                            # the first start date of Task
                            start_date = changed_date
                            # new work found. reset the finish date 
                            finish_date = None
                        elif field_revision[field_name].new_value == 'Done': 
                            # the last finish date of Product Backlog Item or Task
                            finish_date = changed_date

    return (create_date if start_date is None else start_date), finish_date


def compose_item_url(team_context, item_id): 
    item_url_template = "{ORG_URL}/{PROJECT}/_backlogs/backlog/{TEAM}/Backlog items/?workitem={ID}"
    return item_url_template.replace("{ORG_URL}", organization_url).replace("{PROJECT}", team_context.project).replace("{TEAM}", team_context.team).replace("{ID}", str(item_id))


def write_pbi_to_workbook(work_item_list, worksheet, iteration_due_date, append_only): 
    # prepare the excel file 
    ws = worksheet

    export_field_names = copy.deepcopy(pbi_field_names)
    export_field_names.append('Excel.Operation') 
    export_field_excel_operation_index = len(export_field_names)
    export_field_names.append('Excel.Region')
    export_field_excel_region_index = len(export_field_names)
    export_field_names.append('Excel.Planned')
    export_field_excel_planned_index = len(export_field_names)

    export_field_names.append('Excel.ItemUrl') # Item URL to DevOps
    export_field_excel_itemUrl_index = len(export_field_names)

    export_field_names.append('Export.Timestamp') # Export.DueDate
    export_field_export_timestamp_index = len(export_field_names)
    if isinstance(iteration_due_date, datetime): 
        export_timestamp = iteration_due_date.strftime("%Y-%m-%d")

    start_row = 0
    if append_only == False:
        # write the headers to the workbook
        for f in range(len(export_field_names)): 
            simple_name = export_field_names[f].split('.')[-1]
            ws.cell(row=1, column=f+1, value=simple_name)
        start_row = 2
    else:
        start_row = ws.max_row + 1

    # write data to the workbook 
    for i in range(len(work_item_list)): 
        work_item = work_item_list[i]
        r = i + start_row
        start_date = None
        finish_date = None
        for f in range(len(export_field_names)):
            field_name = export_field_names[f] 
            field_value = None
            if field_name == 'Export.Timestamp':
                ws.cell(row=r, column=export_field_export_timestamp_index, value=export_timestamp)
            elif field_name in work_item.fields.keys():
                fc = export_field_names.index(field_name)
                # set the cell value
                field_value = work_item.fields[field_name]
                # further process the cell value
                if field_name == 'System.Id': 
                    ws.cell(row=r, column=export_field_excel_itemUrl_index, value=compose_item_url(team_context, field_value))
                    start_date, finish_date = get_lead_duration(team_context, field_value)
                elif field_name == 'System.Tags': 
                    field_value = field_value.upper()
                    # parse the tags for operation
                    if field_value.find('ELECTRIC') > -1: 
                        ws.cell(row=r, column=export_field_excel_operation_index, value='Eletric')
                    elif field_value.find('GAS') > -1: 
                        ws.cell(row=r, column=export_field_excel_operation_index, value='Gas')
                    else: # count as both 
                        ws.cell(row=r, column=export_field_excel_operation_index, value='Both')
                    # parse the tags for region
                    if field_value.find('INOH') > -1: 
                        ws.cell(row=r, column=export_field_excel_region_index, value='INOH')
                    # parse the tags for Planned or Unplanned 
                    if field_value.find('UNPLANNED') > -1: 
                        ws.cell(row=r, column=export_field_excel_planned_index, value='Unplanned')
                    else: 
                        ws.cell(row=r, column=export_field_excel_planned_index, value='Planned')
                elif field_name == 'System.AssignedTo': 
                    # simplify the AssignedTo object 
                    field_value = work_item.fields["System.AssignedTo"]['displayName']
                elif field_name in ['System.CreatedDate']: 
                    # set the start working date
                    field_value = start_date
                elif field_name in ['System.ChangedDate']:
                    # set the finish working date
                    field_value = finish_date
                ws.cell(row=r, column=fc+1, value=field_value)

    return ws


def write_pbi_to_excel(work_item_list, file_path, sheet_name, iteration_due_date, append_only=False):
    ws = None
    if append_only == True:
        # load the excel file
        wb = load_workbook(file_path)
    else: 
        # prepare the excel file 
        wb = Workbook()
    # get the active sheet
    ws = wb.active
    ws.title = sheet_name

    write_pbi_to_workbook(work_item_list, ws, iteration_due_date, append_only)

    # save to a given file
    wb.save(file_path)    


def get_capacities(team_context, iteration_id, iteration_path): 
    # get capacties for each developer
    work_client = connection.clients.get_work_client()

    capacity_list = []

    get_capacities_response = work_client.get_capacities_with_identity_ref(team_context, iteration_id)
    if get_capacities_response is not None: 
        for dev_capacity in get_capacities_response: 
            team_member = dev_capacity.team_member.display_name
            capacity_per_day = 0
            for dev_activity in dev_capacity.activities: 
                capacity_per_day += dev_activity.capacity_per_day
            capacity_list.append({
                'team_member': team_member, 
                'IterationPath': iteration_path, 
                'days_per_iteration': DAYS_PER_ITERATION, 
                'capacity_per_day': capacity_per_day
            })
    return capacity_list


def write_capacity_to_workbook(capacity_list, worksheet, iteration_due_date, append_only): 
    # prepare the excel file 
    ws = worksheet

    export_field_names = copy.deepcopy(capacity_field_names)

    export_field_names.append('Export.Timestamp') # Export.DueDate
    export_field_export_timestamp_index = len(export_field_names)
    if isinstance(iteration_due_date, datetime): 
        export_timestamp = iteration_due_date.strftime("%Y-%m-%d")

    start_row = 0
    if append_only == False:
        # write the headers to the workbook
        for f in range(len(export_field_names)): 
            simple_name = export_field_names[f].split('.')[-1]
            ws.cell(row=1, column=f+1, value=simple_name)
        start_row = 2
    else:
        start_row = ws.max_row + 1

    # write data to the workbook 
    for i in range(len(capacity_list)): 
        dev_capacity = capacity_list[i]
        r = i + start_row
        start_date = None
        finish_date = None
        for f in range(len(export_field_names)):
            field_name = export_field_names[f] 
            field_value = None
            if field_name == 'Export.Timestamp':
                ws.cell(row=r, column=export_field_export_timestamp_index, value=export_timestamp)
            elif field_name in dev_capacity.keys():
                fc = export_field_names.index(field_name)
                # set the cell value
                field_value = dev_capacity[field_name]
                ws.cell(row=r, column=fc+1, value=field_value)

    return ws


def write_capacity_to_excel(capacity_list, file_path, sheet_name, iteration_due_date, append_only=False):
    ws = None
    if append_only == True:
        # load the excel file
        wb = load_workbook(file_path)
    else: 
        # prepare the excel file 
        wb = Workbook()
    # get the active sheet
    ws = wb.active
    ws.title = sheet_name

    write_capacity_to_workbook(capacity_list, ws, iteration_due_date, append_only)

    # save to a given file
    wb.save(file_path)    


if __name__ == "__main__": 
    parser = argparse.ArgumentParser(description='Retrieve work items from Azure DevOps for a given or current iteration.')
    parser.add_argument('-p', '--project', metavar='<Project Name>', default='CNP.GIS',
                        help='ex. CNP.GIS')
    parser.add_argument('-t', '--team', metavar='<Team Name>', default='CNP.GIS Team',
                        help='ex. CNP.GIS Team')
    parser.add_argument('-i', '--iteration', metavar="<Iteration Path>", 
                        help='ex. CNP.GIS\\Sprint 21.03-A')

    args = parser.parse_args()

    # set the team context 
    team_context = workModels.TeamContext(project=args.project, team=args.team)

    iteration_id = None
    iteration_path = None
    iteration_list = []

    pbi_file_name = None
    capacity_file_name = None
    append_only = False

    if args.iteration is None: 
        print("****** Getting the current iteration ...")
        iteration_id, iteration_path, iteration_due_date = get_current_iteration(team_context)
        iteration_list.append({
            'iteration_id': iteration_id, 
            'iteration_path': iteration_path, 
            'iteration_due_date': iteration_due_date
        })
    elif args.iteration == 'ALL': 
        iteration_list = get_past_iterations(team_context)
        pbi_file_name = 'CNP.GIS_2021_Sprint_{0}_Combined.xlsx'.format('PBI')
        capacity_file_name = 'CNP.GIS_2021_Sprint_{0}_Combined.xlsx'.format('Capacity')
    else: 
        iteration_path = args.iteration
        iteration_id, iteration_due_date = get_iteration(team_context, iteration_path)
        iteration_list.append({
            'iteration_id': iteration_id, 
            'iteration_path': iteration_path, 
            'iteration_due_date': iteration_due_date
        })

    for i in iteration_list: 
        iteration_id = i['iteration_id']
        iteration_path = i['iteration_path']
        iteration_due_date = i['iteration_due_date']

        print("****** Retrieving PBIs for {0} ....".format(iteration_path))
        work_item_list = retrieve_PBIs(team_context, iteration_path)

        local_folder = os.getcwd()
        if pbi_file_name is None: 
            pbi_file_name = iteration_path.replace('\\', '_') + "_PBI.xlsx"
        file_path = os.path.join(os.path.join(local_folder, r"iterations"), pbi_file_name)

        if append_only == False and os.path.exists(file_path):
            raise Exception("File ({0}) already exists.".format(file_path))

        print("****** Storing work items to an Excel file {0} ....".format(file_path)) 
        if iteration_due_date is None:
            iteration_due_date = datetime.now()

        write_pbi_to_excel(work_item_list, file_path, "current_iteration", iteration_due_date, append_only)

        print("****** Retrieving Capacities for {0} ....".format(iteration_path))
        capacity_list = get_capacities(team_context, iteration_id, iteration_path)

        if capacity_file_name is None: 
            capacity_file_name = iteration_path.replace('\\', '_') + "_Capacity.xlsx"
        file_path = os.path.join(os.path.join(local_folder, r"iterations"), capacity_file_name)

        if append_only == False and os.path.exists(file_path):
            raise Exception("File ({0}) already exists.".format(file_path))

        print("****** Storing capacities to an Excel file {0} ....".format(file_path)) 
        if iteration_due_date is None:
            iteration_due_date = datetime.now()

        write_capacity_to_excel(capacity_list, file_path, "capacity", iteration_due_date, append_only)

        append_only = True

    print("****** Completed")
